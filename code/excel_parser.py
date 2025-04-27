# -*- coding: utf-8 -*-
import json
import pandas as pd
import openpyxl
import os
from datetime import datetime
from typing import Dict, Any, List

class ExcelParser:
    """用于解析 Excel 文件并将其转换为 JSON 格式的类，仅提取表格数据。"""

    def __init__(self, file_path: str, doc_type: str = "excel"):
        """
        初始化 ExcelParser 类。
        :param file_path: Excel 文件路径
        :param doc_type: 文档类型
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        self.doc_type = doc_type
        self.result = {
            "doc_type": self.doc_type,
            "file_name": self.file_name,
            "tables": []
        }

    def fill_merged_cells(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> List[List[Any]]:
        """
        填充合并单元格的值到所有相关单元格。
        :param worksheet: openpyxl 的 worksheet 对象
        :return: 填充后的二维数据列表
        """
        max_row = worksheet.max_row
        max_column = worksheet.max_column
        
        data = [[None for _ in range(max_column)] for _ in range(max_row)]
        
        for merged_range in worksheet.merged_cells.ranges:
            top_left_cell = worksheet[merged_range.start_cell.coordinate]
            value = top_left_cell.value
            if isinstance(value, datetime):
                value = value.isoformat()
            elif value is not None:
                value = str(value)
            
            for row in range(merged_range.min_row - 1, merged_range.max_row):
                for col in range(merged_range.min_col - 1, merged_range.max_col):
                    data[row][col] = value
        
        for row in range(max_row):
            for col in range(max_column):
                cell = worksheet.cell(row + 1, col + 1)
                if data[row][col] is None:
                    value = cell.value
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    elif value is not None:
                        value = str(value)
                    data[row][col] = value if value is not None else ""
        
        return data

    def parse(self) -> Dict[str, Any]:
        """
        解析 Excel 文件，提取表格数据，处理合并单元格。
        :return: 解析后的 JSON 数据，仅包含表格
        """
        try:
            wb = openpyxl.load_workbook(self.file_path)

            for sheet_name in wb.sheetnames:
                worksheet = wb[sheet_name]
                
                filled_data = self.fill_merged_cells(worksheet)
                
                if not filled_data:
                    continue
                
                # 过滤空白行
                filled_data = [row for row in filled_data if any(cell != "" for cell in row)]
                if not filled_data:
                    continue
                
                # 假设第一行为表头
                headers = filled_data[0]
                header_counts = {}
                new_headers = []
                for i, header in enumerate(headers):
                    header = str(header).strip() if header else f"Column_{i+1}"
                    if header in header_counts:
                        header_counts[header] += 1
                        new_headers.append(f"{header}_{header_counts[header]}")
                    else:
                        header_counts[header] = 0
                        new_headers.append(header)
                
                df_data = filled_data[1:]
                df = pd.DataFrame(df_data, columns=new_headers)
                
                csv_content = df.to_csv(index=False, encoding='utf-8')
                rows = df.to_dict(orient="records")
                
                self.result["tables"].append({
                    "sheet": sheet_name,
                    "data": csv_content,
                    "rows": rows
                })

        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")

        return self.result

    def split_into_chunks_by_rows(self, table: Dict[str, Any], target_char_limit: int = 60000) -> List[Dict[str, Any]]:
        """
        将 table 的 rows 均分成多个 chunk，使每个 chunk 的字符数接近 target_char_limit。
        :param table: 单个 table 字典，包含 sheet, data, rows
        :param target_char_limit: 目标字符数限制，默认为 60000
        :return: 分割后的 table 列表
        """
        sheet_name = table["sheet"]
        rows = table["rows"]
        if not rows:
            return [table]

        # 估算每行的平均字符数
        sample_size = min(10, len(rows))  # 采样前 10 行（或全部行如果少于 10 行）
        sample_rows = rows[:sample_size]
        avg_row_char_count = 0
        if sample_rows:
            sample_json = json.dumps(sample_rows, ensure_ascii=False)
            avg_row_char_count = len(sample_json) / len(sample_rows)

        # 估算基础 JSON 结构（不含 rows）的字符数
        base_json = {
            "doc_type": self.doc_type,
            "file_name": self.file_name,
            "tables": [
                {
                    "sheet": sheet_name,
                    "data": "",  # 占位符
                    "rows": []
                }
            ]
        }
        base_char_count = len(json.dumps(base_json, ensure_ascii=False))

        # 估算总字符数
        total_row_char_count = avg_row_char_count * len(rows)
        total_char_count = base_char_count + total_row_char_count

        # 计算需要的 chunk 数量
        num_chunks = max(1, int(total_char_count / target_char_limit) + 1)
        rows_per_chunk = max(1, len(rows) // num_chunks)

        # 调整 rows_per_chunk 使字符数更接近 target_char_limit
        while True:
            estimated_char_count = base_char_count + (rows_per_chunk * avg_row_char_count)
            if estimated_char_count <= target_char_limit or rows_per_chunk <= 1:
                break
            num_chunks += 1
            rows_per_chunk = max(1, len(rows) // num_chunks)

        # 分割 rows
        chunks = []
        for i in range(0, len(rows), rows_per_chunk):
            chunk_rows = rows[i:i + rows_per_chunk]
            # 重新生成 data (CSV)
            df = pd.DataFrame(chunk_rows)
            csv_content = df.to_csv(index=False, encoding='utf-8')
            chunks.append({
                "sheet": sheet_name,
                "data": csv_content,
                "rows": chunk_rows
            })

        return chunks

    def save_sheets_to_files(self, output_dir: str = "output", target_char_limit: int = 60000):
        """
        将每个 sheet 的数据保存为单独的 JSON 文件。
        - 如果需要分块，文件名为“excel文件名_sheet名称数字_0.json”。
        - 如果不需要分块，文件名为“excel文件名_sheet名称_0.json”。
        :param output_dir: 输出目录
        :param target_char_limit: 目标字符数限制，默认为 60000
        """
        # 确保输出目录存在
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # 遍历每个 sheet
        for table in self.result["tables"]:
            sheet_name = table["sheet"]
            # 构造基础文件名：excel文件名_sheet名称
            base_name = os.path.splitext(self.file_name)[0]  # 去掉扩展名

            # 检查 JSON 字符数
            sheet_json = {
                "doc_type": self.doc_type,
                "file_name": self.file_name,
                "tables": [table]
            }
            json_str = json.dumps(sheet_json, ensure_ascii=False)
            char_count = len(json_str)

            if char_count <= target_char_limit:
                # 如果字符数未超过限制，直接保存
                output_file_name = f"{base_name}_{sheet_name}_0.json"  # 修改为期望格式
                output_path = os.path.join(output_dir, output_file_name)
                with open(output_path, "w", encoding="utf-8") as f:
                    json.dump(sheet_json, f, ensure_ascii=False, indent=2)
                print(f"Saved sheet '{sheet_name}' to: {output_path}")
            else:
                # 如果字符数超过限制，按行数均分保存
                chunks = self.split_into_chunks_by_rows(table, target_char_limit)
                for i, chunk in enumerate(chunks, 1):  # 从 1 开始编号
                    output_file_name = f"{base_name}_{sheet_name}{i}_0.json"
                    output_path = os.path.join(output_dir, output_file_name)
                    chunk_json = {
                        "doc_type": self.doc_type,
                        "file_name": self.file_name,
                        "tables": [chunk]
                    }
                    with open(output_path, "w", encoding="utf-8") as f:
                        json.dump(chunk_json, f, ensure_ascii=False, indent=2)
                    print(f"Saved chunk {i} of sheet '{sheet_name}' to: {output_path}")

def main():
    try:
        # 1. 使用 ExcelParser 解析 Excel 文件
        file_path = r"C:\Users\dreame\Desktop\电子元件RAG\数据表格纯文字+复杂图片\电子元器件规格归一V02-20240628.xlsx"
        parser = ExcelParser(file_path)
        parsed_json = parser.parse()

        # 调试：检查解析结果
        print("=== Debugging Parsed JSON Data ===")
        print(f"Total sheets: {len(parsed_json['tables'])}")
        for i, table in enumerate(parsed_json['tables']):
            print(f"Sheet {i + 1} - Name: {table['sheet']}")
            print(f"Number of rows: {len(table['rows'])}")
            print(f"Data length: {len(table['data'])} characters")

        # 2. 将每个 sheet 保存为单独的 JSON 文件
        parser.save_sheets_to_files(output_dir="output_test")
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()